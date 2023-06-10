from openpyxl import load_workbook
from time import sleep
import msvcrt
import os
opcion = 0
data = []
wb = load_workbook(filename='DB.xlsx')
hoja = wb.active
contador = 0

def agregar(datos:list):
  for dato in datos:
    hoja.append(dato)
  wb.save('DB.xlsx')
  return "cargado"

def listado():
  columnas = hoja.max_column
  dt = []
  for row in hoja:
    df = []
    for a in range(0,columnas):
      lista = row[a].value
      df.append(lista)
    dt.append(df)
  for i in dt:
    print(i)
  return msvcrt.getche()

def busqueda(titulo):
  columna = hoja.max_column
  for row in hoja:
    lista = row[0].value
    if lista == titulo:
      print(f"{row[0].value} {row[1].value} {row[2].value} {row[3].value}")
    else:
      continue
  return msvcrt.getche()

def eliminar(id):
  hoja.delete_rows(id+1)
  wb.save('DB.xlsx')
  print(f'el libro con id {id} fue eliminado')
  return msvcrt.getche()
while opcion != 5:
  opciones = ['','Agregar Libro','Ver lista de libros','buscar un libro por titulo','Eliminar un libro','salir']
  for i in range(1,len(opciones)):
    print(f'{i}.{opciones[i]}')
  opcion = int(input('ingrese la opcion: '))
  os.system('cls')

  if opcion == 1:
    datos = []
    id = tuple(str(contador))
    titulo = tuple(input('ingrese el titulo: ').split("  "))
    autor = tuple(input('ingrese el nombre del autor: ').split("  "))
    año = tuple(input('ingrese el año de edicion: ').split(" "))
    genero = tuple(input('ingrese el genero: ').split(" "))
    datos.append(tuple(id + titulo + autor + año + genero))
    print(agregar(datos))
    sleep(1)
    os.system('cls')
    contador += 1
  elif opcion == 2:
    print(listado())
    os.system('cls')
  elif opcion == 3:
    print(busqueda(input('ingrese el nombre del titulo que desea buscar: ')))
    os.system('cls')
  elif opcion == 4:
    print(eliminar(int(input('ingrese la ID que desea borrar: '))))
    os.system('cls')
else:
  print('Programa cerrado, precione enter para cerrar esta pestañá')
  msvcrt.getche()
  
    
    
    
    

